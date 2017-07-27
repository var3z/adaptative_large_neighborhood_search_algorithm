from madridALNS import *
from pyomo.opt import *
from time import *
from numpy import *
from math import *
from openpyxl import load_workbook
import winsound


"""
Construccion del modelo
"""

to = clock()
instance = model.create('madridALNS.dat')
tf = clock()
tiempo = tf - to
minutos = tiempo // 60
segundos = int(tiempo - minutos * 60)
print('Tiempo creando el modelo: ' + str(int(minutos)) + ' minutos ' + str(segundos) + ' segundos')


"""
Funcion para resolver modelo
"""
solver='gurobi' #Debe estar instalado

def solveFO(instance):
    opt = SolverFactory(str(solver))
    results = opt.solve(instance, tee=True)
    instance.load(results)
    return float(results.Solution.Objective.__default_objective__['value'])

"""
Datos, variables, parametros y constantes del algoritmo ALNS
"""

# Datos
lineas = ('1_C1', '2_C2', '3_C3', '4_C4', '5_C5', '6_C8', '7_C10')
headways = (3, 4, 5, 6, 10, 12, 15, 20, 30)
nvmax = 10
nvmin = 0
hdmax = headways[-1]
hdmin = headways[0]

# Contadores
it_total = 150      # Numero de iteraciones maximas permitidas
repmax = 50         # Numero maximo que se repite la misma solucion antes de parar
tmax = 2700         # tiempo maximo para iterar en segundos
it_seg = 4          # Numero de iteraciones para actualizar la matriz P(op)
calentamiento = 2   # +2 = Numero de rondas iniciales que no actualiza la matriz P(op)

# Parametros algoritmo
fc = 0.25               # factor de escala para las sigmas

sigma1 = 0.200 * fc     # Puntuacion peor sol. admitida
sigma2 = 0.300 * fc     # Puntuacion sol. ref. nueva
sigma3 = 0.400 * fc     # Puntuacion sol. best. nueva
C = 0.995
contador = 0            # Iteraciones realizadas
contador_puntuacion = 0 # Contador actualizacion puntuacion
mu = 0.9

# Variables del algoritmo
probabilidad = ones(28) / 28  # Probabilidad segun num. operadores
puntos = [0] * 28
veces_oper = [0] * 28
mejorsolucion = []
buvagones = [0] * len(lineas)
buheadways = [0] * len(lineas)

"""
Salida de datos en excel
"""

archivo='plantilla.xlsx' #fichero preparado para guardar resultados de las distintas veces que se ejecute el algoritmo
fil0=3  #fila 3
col0=66 # columna ASCII B
wb = load_workbook(filename = archivo)
ws = wb.active

# funcion escritura datos en excel
def escribe_excel():
    fila=fil0+contador
    col=col0
    alns1 = alns2 = 0
    for dato in aux:
        casilla = str(chr(col) + str(fila))
        ws[casilla] = dato
        col+=1
    for linea in lineas:
        casilla = str(chr(col) + str(fila))
        ws[casilla] = instance.hd[linea].value
        col+=1
    for linea in lineas:
        casilla = str(chr(col) + str(fila))
        ws[casilla] = instance.nv[linea].value
        col += 1
    col1=col2=col0-1
    for proba in probabilidad:
        casilla = str(chr(col1) + chr(col2) + str(fila))
        ws[casilla] = proba
        col2 += 1
        if col2 >  90:
            col1 += 1
            col2 = col0-1


"""
Calculo de solucion inicial Sref
#    Ej: Todas las lineas con 5 vagones y 12 minutos entre trenes
*** La configuracion se fija en MadridALNS.dat
"""
t0 = clock()
solref = solveFO(instance)
tf = clock()

#1a  salida de datos en excel
T = solref
bestsol = solref
sol = solref
aceptada = '-'
vble = 'sol. inicial'
sentido = 'n/a'
time = tf-t0
porcentaje = 0
aux = [contador, T, sol, aceptada, solref, bestsol, vble, sentido, sentido, time, porcentaje]
# el 'sentido' x2 esta puesto para no crear mas vbles
mejorsolucion.append(bestsol)
escribe_excel()


"""
OPERADORES ALNS
"""

# Incremento de headway
def incr_hd(instance, nl):
    posibles = [0] * len(lineas)  # lineas no atacadas
    # Bucle para intentar atacar las 'nl' lineas
    while posibles.count(0) > 0:
        i = random.randint(0, len(lineas))
        hdli = instance.hd[lineas[i]].value
        j = headways.index(instance.hd[lineas[i]].value)
        if posibles[i] == 0:
            if hdli < hdmax:
                instance.hd[lineas[i]].value = headways[j + 1]
            posibles[i] = 1
        if posibles.count(1) == nl:
            break
    instance.preprocess()
    return instance

# Deremento de headway
def decr_hd(instance, nl):
    posibles = [0] * len(lineas)  # lineas no atacadas
    # Bucle para intentar atacar las 'nl' lineas
    while posibles.count(0) > 0:
        i = random.randint(0, len(lineas))
        hdli = instance.hd[lineas[i]].value
        j = headways.index(instance.hd[lineas[i]].value)
        if posibles[i] == 0:
            if hdli > hdmin:
                instance.hd[lineas[i]].value = headways[j - 1]
            posibles[i] = 1
        if posibles.count(1) == nl:
            break
    instance.preprocess()
    return instance

# Incremento de vagones
def incr_vag(instance, nl):
    posibles = [0] * len(lineas)   # lineas no atacadas
    # Bucle para intentar atacar las 'nl' lineas
    while posibles.count(0) > 0:
        i = random.randint(0, len(lineas))
        nvli = instance.nv[lineas[i]].value
        if posibles[i] == 0:
            if nvli < nvmax:
                instance.nv[lineas[i]].value += 1
            posibles[i] = 1
        if posibles.count(1) == nl:
            break
    instance.preprocess()
    return instance

# Deremento de vagones
def decr_vag(instance, nl):
    posibles = [0] * len(lineas)   # lineas no atacadas
    # Bucle para intentar atacar las 'nl' lineas
    while posibles.count(0) > 0:
        i = random.randint(0, len(lineas))
        nvli = instance.nv[lineas[i]].value
        if posibles[i] == 0:
            if nvli > nvmin:
                instance.nv[lineas[i]].value -= 1
            posibles[i] = 1
        if posibles.count(1) == nl:
            break
    instance.preprocess()
    return instance

"""
ALGORITMO ALNS
"""
talgo = clock()
j = 0
operadores = range(len(probabilidad))
while it_total > 0:
    op = int(random.choice(range(len(probabilidad)), 1, replace=False, p=probabilidad))
    veces_oper[op] += 1

    # Llamada a incrementar headway
    if 0 <= op <= len(lineas) - 1:
        for i in range(len(lineas)):
             buheadways[i] = instance.hd[lineas[i]].value # Backup de la configuracion original
        vble = 'headway'     # Vble atacada
        sentido = 'Incr'        # Incremento o decremento
        nl = op + 1
        instance = incr_hd(instance, nl)

    # Llamada a decrementar headway
    elif len(lineas) <= op <= 2 * len(lineas) - 1:
        for i in range(len(lineas)):
             buheadways[i] = instance.hd[lineas[i]].value
        vble = 'headway'
        sentido = 'Decr'
        nl = op - len(lineas) + 1
        instance = decr_hd(instance, nl)

    # Llamada a incrementar vagones
    elif 2 * len(lineas) <=op <= 3 * len(lineas) - 1:
        for i in range(len(lineas)):
             buvagones[i] = instance.nv[lineas[i]].value
        vble = 'vagones'
        sentido = 'Incr'
        nl = op - 2 * len(lineas) - 1
        instance = incr_vag(instance, nl)

    # Llamada a decrementar vagones
    elif 3 * len(lineas) <= op <= 4 * len(lineas) - 1:
        for i in range(len(lineas)):
             buvagones[i] = instance.nv[lineas[i]].value
        vble = 'vagones'
        sentido = 'Decr'
        nl = op - 3 * len(lineas) - 1
        instance = decr_vag(instance, nl)

# Llamada a solver con nueva configuracion recibida del operador
    t0 = clock()
    print probabilidad
    sol = solveFO(instance)
    contador += 1
    if contador >= calentamiento:
        contador_puntuacion += 1
    T = C * T
    porcentaje = (1 - sol / bestsol) * 100
    aceptada ='-'

# Comprueba soluciones y mejora la probabilidad de volver a aplicar dicho operador si procede
    if porcentaje > 0 :
        bestsol = sol
        puntos[op] += sigma3
    elif sol < solref:
        solref = sol
        puntos[op] += sigma2
    else:
        acepta = (1 - exp(- T / (sol - bestsol + 0.001))) #*C
        soladmisible = random.choice([0, 1], 1, replace=False, p=[1 - acepta, acepta])
        if soladmisible == 1:
            puntos[op] += sigma1
            solref = sol
            aceptada = 'SI'
        # Corrige las modificaciones sobre el modelo, ya que la solucion no es mejor que la de referencia:
        else:
            aceptada = 'NO'
            if 0 <= op <= (2 * len(lineas) - 1):
                for i in range(len(lineas)):
                    instance.hd[lineas[i]].value = buheadways[i]
            else:
                for i in range(len(lineas)):
                    instance.nv[lineas[i]].value = buvagones[i]
            instance.preprocess()

    # Actualizacion de las probabilidades de los operadores
    if contador >= calentamiento and contador_puntuacion >= it_seg:
        for i in operadores:
            if veces_oper[i] != 0:
                probabilidad[i] = (1 - mu) * probabilidad[i] + mu * (puntos[i] / veces_oper[i])
                puntos[i] = veces_oper[i] = 0
        contador_puntuacion = 0
    probabilidad = probabilidad / sum(p for p in probabilidad)
    mejorsolucion.append(bestsol)
    tf = clock()
    aux = [contador, T, sol, aceptada, solref, bestsol, vble, sentido, op, time, porcentaje]
    print aux
    escribe_excel()
    # Bucle para terminar si encuentra n veces seguidas la misma solucion optima
    if len(mejorsolucion) > 1:
        if mejorsolucion[-1] == mejorsolucion[-2]:
            j += 1
            if j == repmax:
                break
        else:
            j = 0
    if clock() - talgo > tmax:
        break
    it_total -= 1
tfin = clock()
ttotal = tfin - talgo
minutos = int(ttotal // 60)
segundos = round(ttotal - 60 * minutos,1)
ws['A1'] = 't = '+str(minutos)+ '\' ' + str(segundos)+ '\"'
nombrearchivo = strftime('%d%m%y - %H,%M')
wb.save(nombrearchivo+'.xlsx')
# Aviso sonoro de que ha terminado
winsound.Beep(1000,250)
winsound.Beep(37,1)
winsound.Beep(1000,750)